import praw
import random
import os
import openpyxl

def Pokeformat(name,rownum):
    HP = sheet.cell(row=rownum, column=2).value
    HP = int(HP)
    HP = str(HP)

    ATK = sheet.cell(row=rownum, column=3).value
    ATK = int(ATK)
    ATK = str(ATK)

    DEF = sheet.cell(row=rownum, column=4).value
    DEF = int(DEF)
    DEF = str(DEF)

    SATK = sheet.cell(row=rownum, column=5).value
    SATK = int(SATK)
    SATK = str(SATK)

    SDEF = sheet.cell(row=rownum, column=6).value
    SDEF = int(SDEF)
    SDEF = str(SDEF)

    SPD = sheet.cell(row=rownum, column=7).value
    SPD = int(SPD)
    SPD = str(SPD)

    Total = sheet.cell(row=rownum, column=8).value
    Total = int(Total)
    Total = str(Total)


    comment = ("**"+name+"**\n\n"

                      "* HP: "+HP+"\n\n"

                      "* ATK: "+ATK+"\n\n"

                      "* DEF: "+DEF+"\n\n"

                      "* SATK: "+SATK+"\n\n"

                      "* SDEF: "+SDEF+"\n\n"

                      "* SPD: "+SPD+"\n\n"

                      "* Total: "+Total+"\n\n"
               )
    return comment


wb = openpyxl.load_workbook('Pokemon.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

REDDIT_USERNAME = 'SagaTracker'  # YOUR USERNAME as string
REDDIT_PASS = 'mudkip123'  # YOUR PASSWORD as string
r = praw.Reddit('Pokereply 0.1')
r.login(REDDIT_USERNAME, REDDIT_PASS)
subreddit = r.get_subreddit('345gbvwe5') #Name of subreddit for the bot to work in


if not os.path.isfile("posts_saved.txt"):
    posts_saved = []

else:
    # Read the file into a list and remove any empty values
    with open("posts_saved.txt", "r") as f:
        posts_saved = f.read()
        posts_saved = posts_saved.split("\n")
        posts_saved = filter(None, posts_saved)

while True:
    subreddit_comments = subreddit.get_comments()
    for comment in subreddit_comments:
        if "!Stats" in comment.body and comment.id not in posts_saved:
            #Known problem here, if more than one pokemon name is in the comment, the bot finds both
            for rowNum in range(2, sheet.max_row ):
                PokemonName = sheet.cell(row=rowNum, column=1).value
                if PokemonName in comment.body:
                    comment.reply(Pokeformat(PokemonName, rowNum))

                posts_saved.append(comment.id)






